"""
conftest.py - Fixtures compartidos para todos los tests
"""

import pytest
import json
import tempfile
import shutil
from pathlib import Path
from unittest.mock import Mock, MagicMock
import numpy as np
from PIL import Image

from src.data_models import EmbeddingRecord


@pytest.fixture(scope="session")
def temp_test_dir():
    """Directorio temporal para todos los tests de la sesión"""
    temp_dir = tempfile.mkdtemp(prefix="multidoc_test_")
    yield Path(temp_dir)
    shutil.rmtree(temp_dir)


@pytest.fixture
def mock_excel_file(temp_test_dir):
    """Crea un archivo Excel de prueba simple"""
    import openpyxl
    
    excel_path = temp_test_dir / "test_data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Datos de prueba (10x10)
    for r in range(1, 11):
        for c in range(1, 11):
            ws.cell(row=r, column=c, value=f"R{r}C{c}")
    
    # Agregar algunos números
    ws['A1'] = "Bodega"
    ws['B1'] = "Tonelaje"
    ws['A2'] = "B1"
    ws['B2'] = 1234.5
    ws['A3'] = "B2"
    ws['B3'] = 5678.9
    
    wb.save(excel_path)
    return excel_path


@pytest.fixture
def mock_pdf_file(temp_test_dir):
    """Crea un PDF de prueba simple"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    
    pdf_path = temp_test_dir / "test_document.pdf"
    c = canvas.Canvas(str(pdf_path), pagesize=letter)
    c.drawString(100, 750, "Test PDF Document")
    c.drawString(100, 730, "Recalada: R-2024-001")
    c.drawString(100, 710, "Nave: TEST VESSEL")
    c.drawString(100, 690, "Bodega B1: 1234.5 toneladas")
    c.save()
    
    return pdf_path


@pytest.fixture
def mock_image_file(temp_test_dir):
    """Crea una imagen PNG de prueba"""
    img_path = temp_test_dir / "test_chunk_r0_c0.png"
    
    # Crear imagen simple (100x100 blanca con texto simulado)
    img = Image.new('RGB', (100, 100), color='white')
    img.save(img_path)
    
    return img_path


@pytest.fixture
def sample_markdown_text():
    """Texto Markdown de ejemplo (salida de ADE)"""
    return """
# Operación de Nave

**Recalada:** R-2024-001
**Nave:** MV EXAMPLE SHIP
**Fecha Inicio:** 2024-01-15
**Fecha Fin:** 2024-01-16

## Bodegas

| Bodega | Tonelaje | Lotes |
|--------|----------|-------|
| B1     | 1234.5   | L001, L002 |
| B2     | 5678.9   | L003 |

## Clientes

- Cliente A
- Cliente B

## Facturación

- Lote L001: FAC-001, FAC-002
- Lote L002: FAC-003
- Lote L003: FAC-004
"""


@pytest.fixture
def sample_structured_json():
    """JSON estructurado de ejemplo (salida del Paso 4)"""
    return {
        "recalada": "R-2024-001",
        "nave_nombre": "MV EXAMPLE SHIP",
        "fecha_inicio_operacion": "2024-01-15",
        "fecha_fin_operacion": "2024-01-16",
        "bodegas": [
            {
                "bodega": "B1",
                "tonelaje": 1234.5,
                "lotes": ["L001", "L002"]
            },
            {
                "bodega": "B2",
                "tonelaje": 5678.9,
                "lotes": ["L003"]
            }
        ],
        "clientes": ["Cliente A", "Cliente B"],
        "lotes_facturacion": [
            {
                "lote": "L001",
                "cliente": "Cliente A",
                "codigos_facturacion": ["FAC-001", "FAC-002"],
                "bodega": "B1",
                "tonelaje": 617.25
            },
            {
                "lote": "L002",
                "cliente": "Cliente A",
                "codigos_facturacion": ["FAC-003"],
                "bodega": "B1",
                "tonelaje": 617.25
            },
            {
                "lote": "L003",
                "cliente": "Cliente B",
                "codigos_facturacion": ["FAC-004"],
                "bodega": "B2",
                "tonelaje": 5678.9
            }
        ]
    }


@pytest.fixture
def sample_embedding():
    """Embedding de ejemplo (512 dimensiones, normalizado)"""
    embedding = np.random.randn(512).astype(np.float32)
    embedding = embedding / np.linalg.norm(embedding)  # Normalizar
    return embedding


@pytest.fixture
def sample_embedding_record(sample_embedding):
    """EmbeddingRecord completo de ejemplo"""
    return EmbeddingRecord(
        embedding=sample_embedding.tolist(),
        document="Recalada: R-2024-001. Nave: MV EXAMPLE SHIP. Bodega B1: 1234.5 toneladas",
        metadata={
            "type": "text",
            "source_file": "test_document",
            "json_path": "processed/extracted_tables/test_document_structure.json"
        }
    )


@pytest.fixture
def mock_clip_model():
    """Mock del modelo CLIP para evitar carga real en tests"""
    mock_model = MagicMock()
    mock_processor = MagicMock()
    
    # Simular salida del modelo
    mock_features = MagicMock()
    mock_features.cpu.return_value.numpy.return_value = np.random.randn(1, 512).astype(np.float32)
    
    mock_model.get_image_features.return_value = mock_features
    mock_model.get_text_features.return_value = mock_features
    
    return mock_model, mock_processor


@pytest.fixture
def mock_chroma_collection():
    """Mock de una colección de ChromaDB"""
    mock_collection = MagicMock()
    mock_collection.count.return_value = 10
    mock_collection.add = MagicMock()
    mock_collection.query.return_value = {
        "ids": [["text_test_document", "excel_test_data_r0_c0"]],
        "distances": [[0.2, 0.3]],
        "metadatas": [[
            {"type": "text", "source_file": "test_document"},
            {"type": "excel_image", "source_file": "test_data.xlsx", "chunk": "r0_c0"}
        ]],
        "documents": [["Documento de prueba 1", "Documento de prueba 2"]]
    }
    return mock_collection


@pytest.fixture
def mock_llm_response():
    """Respuesta simulada de un LLM"""
    class MockLLMResponse:
        content = """Basándome en los documentos recuperados:

La operación corresponde a la recalada R-2024-001 de la nave MV EXAMPLE SHIP [Documento 1].

El tonelaje total en la bodega B1 fue de 1234.5 toneladas [Documento 1], distribuido en los lotes L001 y L002 [Documento 2].

El cliente atendido fue Cliente A [Documento 1]."""
    
    return MockLLMResponse()


@pytest.fixture(autouse=True)
def reset_singletons():
    """Resetea singletons entre tests para evitar contaminación"""
    # Resetear el singleton de CLIPEncoder
    from src.embeddings.clip_encoder import CLIPEncoder
    CLIPEncoder._instance = None
    
    yield
    
    # Cleanup después del test
    CLIPEncoder._instance = None