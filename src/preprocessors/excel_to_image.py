"""
excel_to_image.py - Conversión de Excel a imagen PNG
(Versión MODIFICADA usando Matplotlib y "Chunking" de 50x50 celdas)
"""

import json
from pathlib import Path
from datetime import datetime
import openpyxl
import matplotlib.pyplot as plt
import logging # Usar logging en lugar de get_logger para evitar error de importación circular
from src.utils.config import INPUT_EXCEL_DIR, EXCEL_IMAGES_DIR

# Configura el logger para este módulo
logger = logging.getLogger(__name__)


def excel_to_image(excel_path: str, sheet_name: str = None, range_to_capture: str = None) -> dict:
    """
    Convierte un rango de Excel en múltiples imágenes PNG,
    divididas en "chunks" (fragmentos) de 50x50 celdas.
    
    Args:
        excel_path: Ruta al archivo Excel
        sheet_name: Nombre de la hoja (default: primera hoja)
        range_to_capture: Rango a capturar (default: auto-detectar)
    
    Returns:
        dict: Información sobre las imágenes generadas (ahora una lista)
    """
    try:
        logger.info(f"Iniciando conversión de Excel (en chunks): {excel_path}")
        
        # Abrir archivo Excel (data_only=True para no ver fórmulas)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active if sheet_name is None else wb[sheet_name]
        
        # Detectar rango automáticamente si no se provee
        if range_to_capture is None:
            range_to_capture = ws.dimensions
        
        logger.info(f"Usando rango total: {range_to_capture}")
        
        # 1. Extraer TODOS los datos del rango a una lista de listas
        data = []
        try:
            for row in ws[range_to_capture]:
                data.append([cell.value if cell.value is not None else "" for cell in row])
        except Exception as e:
            logger.warning(f"No se pudo leer el rango '{range_to_capture}'. Error: {e}")
            return {"status": "error", "message": f"Rango inválido: {range_to_capture}"}

        # 2. Filtrar filas vacías (para tener un 'data' limpio)
        data = [row for row in data if any(str(cell).strip() != "" for cell in row)]
        if not data:
            logger.warning(f"No se encontraron datos en el rango {range_to_capture} para {excel_path}")
            return {"status": "error", "message": "Rango vacío o sin datos"}

        num_rows = len(data)
        num_cols = len(data[0]) if num_rows > 0 else 0
        if num_cols == 0:
             return {"status": "error", "message": "No hay columnas en los datos."}

        # --- ¡NUEVA LÓGICA DE CHUNKING! ---
        
        CHUNK_ROWS = 50
        CHUNK_COLS = 50
        generated_image_paths = [] # Lista para guardar las rutas de las imágenes
        
        file_name_stem = Path(excel_path).stem # Nombre base sin extensión
        
        # Iterar sobre el 'data' en chunks de 50 filas
        for r_start in range(0, num_rows, CHUNK_ROWS):
            # Iterar sobre el 'data' en chunks de 50 columnas
            for c_start in range(0, num_cols, CHUNK_COLS):
                
                r_end = min(r_start + CHUNK_ROWS, num_rows)
                c_end = min(c_start + CHUNK_COLS, num_cols)
                
                # 3. Extraer los datos solo para este chunk
                chunk_data = [row[c_start:c_end] for row in data[r_start:r_end]]
                
                # 4. Validar que el chunk tenga contenido real (no solo celdas vacías)
                if not chunk_data or not any(any(str(cell).strip() for cell in row) for row in chunk_data):
                    continue # Omitir este chunk vacío

                # 5. Lógica de Matplotlib (movida dentro del loop)
                num_chunk_rows = len(chunk_data)
                num_chunk_cols = len(chunk_data[0])
                
                fig_width = max(15, num_chunk_cols * 1.5) 
                fig_height = max(5, num_chunk_rows * 0.35)
                
                fig, ax = plt.subplots(figsize=(fig_width, fig_height))
                ax.axis('off')

                table = ax.table(cellText=chunk_data, loc='center', cellLoc='left')
                table.auto_set_font_size(False)
                table.set_fontsize(8)
                table.scale(1, 1.2) 

                # 6. Definir la ruta de salida ÚNICA para este chunk
                chunk_image_name = f"{file_name_stem}_chunk_r{r_start}_c{c_start}.png"
                image_path = EXCEL_IMAGES_DIR / chunk_image_name
                
                # 7. Guardar la figura del CHUNK
                plt.savefig(image_path, bbox_inches='tight', dpi=200)
                plt.close(fig) # Cerrar figura para liberar memoria
                
                generated_image_paths.append(str(image_path))
        
        # --- FIN DE LA LÓGICA DE CHUNKING ---
        
        if not generated_image_paths:
            logger.warning(f"El rango {range_to_capture} no produjo ningún chunk de imagen con datos.")
            return {"status": "error", "message": "No se generaron chunks de imagen."}

        logger.info(f"Excel procesado y guardado como {len(generated_image_paths)} chunks PNG.")

        # Guardar información de metadata (AHORA CON LA LISTA DE IMÁGENES)
        metadata = {
            "original_file": Path(excel_path).name,
            "image_files": generated_image_paths, # <-- Clave plural
            "sheet_name": ws.title,
            "range": range_to_capture,
            "chunk_size": f"{CHUNK_ROWS}x{CHUNK_COLS}",
            "timestamp": datetime.now().isoformat()
        }
        
        # Guardar el JSON de metadata
        save_excel_metadata(file_name_stem, metadata)
        
        return {
            "status": "success",
            "metadata": metadata
        }
    
    except Exception as e:
        logger.error(f"Error al convertir Excel a imagen: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}


def save_excel_metadata(file_name: str, metadata: dict) -> bool:
    """
    Guarda metadata del Excel procesado en un archivo JSON.
    (Esta función no necesita cambios, ya guarda el 'dict' que le pases)
    """
    try:
        metadata_path = EXCEL_IMAGES_DIR / f"{Path(file_name).stem}_metadata.json"
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
        logger.info(f"Metadata de Excel guardada: {metadata_path.name}")
        return True
    except Exception as e:
        logger.error(f"Error al guardar metadata de Excel: {e}")
        return False


def process_all_excels() -> list:
    """
    Procesa todos los Excel en el directorio de entrada.
    (Esta función no necesita cambios)
    """
    results = []
    excel_files = list(INPUT_EXCEL_DIR.glob("*.xlsx"))
    
    logger.info(f"Encontrados {len(excel_files)} archivos Excel")
    
    for excel_file in excel_files:
        result = excel_to_image(str(excel_file))
        results.append(result)
    
    return results

# --- Código de prueba (para ejecutar este archivo directamente) ---
if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    logger.info("Iniciando prueba de conversión Excel -> PNG (con chunks)...")
    
    INPUT_EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    EXCEL_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    
    test_excel_path = INPUT_EXCEL_DIR / "test_excel_file_large.xlsx"
    
    # Crear un Excel de prueba más grande para probar el chunking
    if not test_excel_path.exists():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "HojaGrande"
            # Llenar 60x60 celdas
            for r in range(1, 61):
                for c in range(1, 61):
                    ws.cell(row=r, column=c, value=f"R{r}C{c}")
            
            # Poner una fórmula en la esquina
            ws['B61'] = "=SUM(B2:B60)"
            wb.save(test_excel_path)
            logger.info(f"Creado archivo Excel de prueba GRANDE: {test_excel_path.name}")
        except Exception as e:
            logger.error(f"No se pudo crear el archivo Excel de prueba: {e}")

    # Ejecutar el procesamiento
    process_results = process_all_excels()
    
    print("\n--- Resultados del Procesamiento ---")
    print(json.dumps(process_results, indent=2))
    
    # Mostrar la nueva metadata
    if process_results and process_results[0].get('status') == 'success':
        metadata = process_results[0].get('metadata', {})
        image_files = metadata.get('image_files', [])
        if image_files:
            logger.info(f"¡Éxito! Se generaron {len(image_files)} imágenes chunk:")
            for img_path in image_files:
                logger.info(f"  - {Path(img_path).name}")
            # Deberías ver 4 chunks: (0,0), (0,50), (50,0), (50,50)